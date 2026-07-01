"""
Exportar/importar el catálogo de productos en Excel (.xlsx) con openpyxl.

    exportar_xlsx() -> bytes         # catálogo completo (base + nuevos + ajustes)
    importar_xlsx(contenido) -> dict # crea/actualiza productos desde un Excel
"""

import io

from openpyxl import Workbook, load_workbook

from agent import catalogo, db

# Orden y encabezados de las columnas del Excel.
COLUMNAS = [
    "codigo", "categoria", "nombre", "descripcion", "material", "uso",
    "tallas", "colores", "precio_publico", "precio_mayoreo", "marca",
    "observaciones", "sin_stock",
]
ENCABEZADOS = {
    "codigo": "Código", "categoria": "Categoría", "nombre": "Nombre",
    "descripcion": "Descripción", "material": "Material", "uso": "Uso",
    "tallas": "Tallas", "colores": "Colores", "precio_publico": "Precio público",
    "precio_mayoreo": "Precio mayoreo", "marca": "Marca",
    "observaciones": "Observaciones", "sin_stock": "Sin stock",
}


def _to_int(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = "".join(ch for ch in str(v) if ch.isdigit())
    return int(s) if s else None


def exportar_xlsx() -> bytes:
    """Genera un .xlsx con todo el catálogo (ordenado por categoría y código)."""
    prods = catalogo.buscar("", incluir_sin_stock=True)
    prods = sorted(prods, key=lambda p: (p.get("categoria", ""), p.get("codigo", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append([ENCABEZADOS[c] for c in COLUMNAS])
    for p in prods:
        fila = []
        for c in COLUMNAS:
            if c == "sin_stock":
                fila.append("SI" if p.get("sin_stock") else "")
            else:
                fila.append(p.get(c, ""))
        ws.append(fila)
    # Anchos legibles.
    for i, _ in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[chr(64 + i)].width = 20
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_SI = {"si", "sí", "x", "1", "true", "yes", "verdadero", "agotado", "sin stock"}


def importar_xlsx(contenido: bytes) -> dict:
    """Crea o actualiza productos desde un Excel. Empareja por 'Código':
    si el código existe -> actualiza (precio, nombre, sin stock); si no -> crea."""
    wb = load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 2:
        return {"error": "El archivo no tiene filas de productos."}

    # Mapear encabezados -> campo (acepta el nombre bonito o el técnico).
    header = [str(h).strip().lower() if h is not None else "" for h in filas[0]]
    col_de = {}
    for c in COLUMNAS:
        etiqueta = ENCABEZADOS[c].lower()
        for j, h in enumerate(header):
            if h == etiqueta or h == c:
                col_de[c] = j
                break
    if "nombre" not in col_de and "codigo" not in col_de:
        return {"error": "No encontré las columnas Código/Nombre en el Excel."}

    existentes = {p["codigo"] for p in catalogo.buscar("", incluir_sin_stock=True)}

    def val(row, campo):
        j = col_de.get(campo)
        return row[j] if j is not None and j < len(row) else None

    creados = actualizados = 0
    errores = []
    for row in filas[1:]:
        nombre = (str(val(row, "nombre")).strip() if val(row, "nombre") else "")
        codigo = (str(val(row, "codigo")).strip().upper() if val(row, "codigo") else "")
        if not nombre and not codigo:
            continue
        pub = _to_int(val(row, "precio_publico"))
        may = _to_int(val(row, "precio_mayoreo"))
        sin = 1 if str(val(row, "sin_stock") or "").strip().lower() in _SI else 0
        obs = val(row, "observaciones")

        if codigo and codigo in existentes:
            campos = {"sin_stock": sin}
            if pub is not None:
                campos["precio_publico"] = pub
            if may is not None:
                campos["precio_mayoreo"] = may
            if nombre:
                campos["nombre"] = nombre
            if obs is not None:
                campos["observaciones"] = str(obs)
            db.set_override(codigo, campos)
            actualizados += 1
        else:
            if not nombre:
                errores.append(f"Fila con código '{codigo}' sin nombre: omitida.")
                continue
            nuevo = db.crear_producto({
                "codigo": codigo,
                "nombre": nombre,
                "categoria": str(val(row, "categoria") or "General"),
                "precio_publico": pub or 0,
                "precio_mayoreo": may or 0,
                "descripcion": str(val(row, "descripcion") or ""),
                "uso": str(val(row, "uso") or ""),
            })
            if sin:
                db.set_override(nuevo, {"sin_stock": 1})
            existentes.add(nuevo)
            creados += 1

    return {"ok": True, "creados": creados, "actualizados": actualizados,
            "errores": errores[:10]}
